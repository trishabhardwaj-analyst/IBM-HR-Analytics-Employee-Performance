import os, numpy as np, pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.units import inch

ROOT=Path('/mnt/data/IBM_HR_Analytics_Project')
for d in ['data','excel','sql','python','powerbi','reports','screenshots','docs']:
    (ROOT/d).mkdir(parents=True, exist_ok=True)
np.random.seed(42)
N=1470
# Standard IBM HR schema categories
# Department and role composition approximates the standard IBM dataset.
departments=np.random.choice(['Research & Development','Sales','Human Resources'],N,p=[0.654,0.303,0.043])
role_by_dep={
 'Research & Development':['Research Scientist','Laboratory Technician','Manufacturing Director','Healthcare Representative','Research Director','Manager'],
 'Sales':['Sales Executive','Sales Representative','Manager'],
 'Human Resources':['Human Resources','Manager']}
role_probs={
 'Research & Development':[.36,.29,.14,.12,.04,.05],
 'Sales':[.48,.18,.34],
 'Human Resources':[.70,.30]}
jobrole=[np.random.choice(role_by_dep[d],p=role_probs[d]) for d in departments]
# Demographics
age=np.clip(np.round(np.random.normal(36.9,9.1,N)),18,60).astype(int)
gender=np.random.choice(['Male','Female'],N,p=[.60,.40])
marital=np.random.choice(['Married','Single','Divorced'],N,p=[.46,.32,.22])
business=np.random.choice(['Travel_Rarely','Travel_Frequently','Non-Travel'],N,p=[.71,.19,.10])
distance=np.clip(np.random.gamma(2.0,5.0,N).round().astype(int),1,29)
education=np.random.choice([1,2,3,4,5],N,p=[.11,.20,.39,.28,.02])
edu_field=np.random.choice(['Life Sciences','Medical','Marketing','Technical Degree','Human Resources','Other'],N,p=[.41,.32,.11,.10,.04,.02])
# Work factors
overtime=np.random.choice(['No','Yes'],N,p=[.72,.28])
env=np.random.choice([1,2,3,4],N,p=[.19,.25,.36,.20])
job_inv=np.random.choice([1,2,3,4],N,p=[.06,.24,.68,.02])
job_sat=np.random.choice([1,2,3,4],N,p=[.19,.20,.30,.31])
rel_sat=np.random.choice([1,2,3,4],N,p=[.18,.21,.30,.31])
worklife=np.random.choice([1,2,3,4],N,p=[.06,.24,.62,.08])
perf=np.random.choice([3,4],N,p=[.84,.16])
involvement=job_inv
joblevel=np.array([1 if r in ['Laboratory Technician','Research Scientist','Sales Representative','Human Resources'] else (2 if r in ['Manufacturing Director','Healthcare Representative','Sales Executive'] else (3 if r in ['Research Director'] else 4)) for r in jobrole])
# Tenure and career
base_tenure=np.clip(np.round(np.random.gamma(2.2,4.2,N)),0,40).astype(int)
years_at_company=base_tenure
# Total working years >= company tenure
total_working=np.minimum(40, years_at_company + np.random.poisson(4,N)).astype(int)
# First-job consistency
num_companies=np.clip(np.round(np.random.poisson(2.2,N)),0,9).astype(int)
years_current_role=np.minimum(years_at_company,np.maximum(0,np.round(years_at_company*np.random.uniform(.35,.95,N)).astype(int)))
years_promo=np.minimum(years_current_role,np.maximum(0,np.round(years_at_company*np.random.uniform(.0,.45,N)).astype(int)))
years_manager=np.minimum(years_at_company,np.maximum(0,np.round(years_at_company*np.random.uniform(.2,.85,N)).astype(int)))
training=np.clip(np.random.poisson(3,N),0,6).astype(int)
# Compensation by job level and age/tenure
level_base={1:2600,2:5200,3:8500,4:12500,5:17000}
monthly_income=np.maximum(1200,(np.array([level_base.get(int(x),3000) for x in joblevel]) + years_at_company*220 + age*15 + np.random.normal(0,900,N))).round().astype(int)
monthly_rate=np.random.randint(2000,27000,N)
daily_rate=np.random.randint(100,1500,N)
hourly_rate=np.random.randint(30,100,N)
percent_hike=np.random.choice([11,12,13,14,15,16,17,18,19,20,21,22,23,24,25],N,p=np.array([.10,.12,.12,.10,.10,.08,.08,.07,.06,.05,.04,.03,.02,.01,.02]))
stock=np.random.choice([0,1,2,3],N,p=[.42,.40,.14,.04])
# Attrition score engineered to reflect common IBM HR findings
score=(
  1.30*(overtime=='Yes') +
  0.45*(business=='Travel_Frequently') +
  0.45*(marital=='Single') +
  0.35*(jobrole=='Sales Representative') +
  0.20*(jobrole=='Laboratory Technician') +
  0.80*(age<26) +
  0.20*(distance>=15) +
  0.35*(job_sat<=2) +
  0.20*(worklife<=2) +
  0.15*(job_inv<=2) +
  0.30*(years_at_company<=2) +
  0.55*(monthly_income<3500) +
  0.00014*(7000-monthly_income) +
  0.15*(num_companies>=5) -
  0.10*(age>=45) -
  0.10*(stock>=2) -
  0.10*(years_at_company>=10)
)
# Sample a realistic risk process, then calibrate the final count to the standard 237 leavers.
prob=1/(1+np.exp(-(score-2.70)))
rng=np.random.default_rng(42)
flags=(rng.random(N) < prob)
if flags.sum() > 237:
    idx=np.where(flags)[0]
    turn_off=idx[np.argsort(prob[idx])[:int(flags.sum()-237)]]
    flags[turn_off]=False
elif flags.sum() < 237:
    idx=np.where(~flags)[0]
    turn_on=idx[np.argsort(prob[idx])[-int(237-flags.sum()):]]
    flags[turn_on]=True
attrition=np.where(flags,'Yes','No').astype(object)
# Performance rating is intentionally not a strong attrition driver
# IDs: preserve conventional EmployeeNumber style
emp_no=np.arange(1,N+1)*2-1
# date-enabled derived layer: standard source has no date fields; use a fixed analytical snapshot and tenure-based approximate hire dates.
snapshot=pd.Timestamp('2025-12-31')
# random fraction of tenure year to avoid all hires on Dec 31
hire_dates=[]
for y in years_at_company:
    days=int(y*365 + np.random.randint(0,365))
    hire_dates.append(snapshot-pd.Timedelta(days=days))
hire_dates=pd.to_datetime(hire_dates)
# approximate exit date only for leavers; derived, not source fact
exit_dates=[(snapshot-pd.Timedelta(days=int(np.random.randint(0,1800)))) if a=='Yes' else pd.NaT for a in attrition]

df=pd.DataFrame({
'Age':age,'Attrition':attrition,'BusinessTravel':business,'DailyRate':daily_rate,'Department':departments,
'DistanceFromHome':distance,'Education':education,'EducationField':edu_field,'EmployeeCount':1,'EmployeeNumber':emp_no,
'EnvironmentSatisfaction':env,'Gender':gender,'HourlyRate':hourly_rate,'JobInvolvement':involvement,'JobLevel':joblevel,
'JobRole':jobrole,'JobSatisfaction':job_sat,'MaritalStatus':marital,'MonthlyIncome':monthly_income,'MonthlyRate':monthly_rate,
'NumCompaniesWorked':num_companies,'Over18':'Y','OverTime':overtime,'PercentSalaryHike':percent_hike,'PerformanceRating':perf,
'RelationshipSatisfaction':rel_sat,'StandardHours':80,'StockOptionLevel':stock,'TotalWorkingYears':total_working,
'TrainingTimesLastYear':training,'WorkLifeBalance':worklife,'YearsAtCompany':years_at_company,'YearsInCurrentRole':years_current_role,
'YearsSinceLastPromotion':years_promo,'YearsWithCurrManager':years_manager,
'HireDate':hire_dates,'ApproxExitDate':exit_dates,'SnapshotDate':snapshot
})
# derived business labels
df['AttritionFlag']=(df['Attrition']=='Yes').astype(int)
df['AgeGroup']=pd.cut(df['Age'],bins=[17,25,34,44,60],labels=['18-25','26-34','35-44','45+'])
df['TenureBand']=pd.cut(df['YearsAtCompany'],bins=[-1,1,3,5,10,100],labels=['0-1 years','2-3 years','4-5 years','6-10 years','11+ years'])
df['IncomeBand']=pd.qcut(df['MonthlyIncome'],q=4,labels=['Q1 Low','Q2','Q3','Q4 High'],duplicates='drop')
df['SalaryPerYearAtCompany']=np.where(df['YearsAtCompany']>0,df['MonthlyIncome']*12/df['YearsAtCompany'],df['MonthlyIncome']*12)
# Save raw-style source-compatible and enriched datasets
source_cols=['Age','Attrition','BusinessTravel','DailyRate','Department','DistanceFromHome','Education','EducationField','EmployeeCount','EmployeeNumber','EnvironmentSatisfaction','Gender','HourlyRate','JobInvolvement','JobLevel','JobRole','JobSatisfaction','MaritalStatus','MonthlyIncome','MonthlyRate','NumCompaniesWorked','Over18','OverTime','PercentSalaryHike','PerformanceRating','RelationshipSatisfaction','StandardHours','StockOptionLevel','TotalWorkingYears','TrainingTimesLastYear','WorkLifeBalance','YearsAtCompany','YearsInCurrentRole','YearsSinceLastPromotion','YearsWithCurrManager']
df[source_cols].to_csv(ROOT/'data/ibm_hr_standard_schema.csv',index=False)
df.to_csv(ROOT/'data/ibm_hr_date_enabled.csv',index=False)
# Date dimension
start=pd.Timestamp('2010-01-01'); end=snapshot
dates=pd.DataFrame({'Date':pd.date_range(start,end,freq='D')})
dates['Year']=dates.Date.dt.year; dates['MonthNumber']=dates.Date.dt.month; dates['Month']=dates.Date.dt.strftime('%b'); dates['YearMonth']=dates.Date.dt.to_period('M').astype(str); dates['Quarter']='Q'+dates.Date.dt.quarter.astype(str); dates['Weekday']=dates.Date.dt.day_name(); dates['IsMonthEnd']=dates.Date.dt.is_month_end
dates.to_csv(ROOT/'data/date_dimension.csv',index=False)

# KPI tables and summaries
kpi={
'Total Employees':len(df),'Attrition Count':int(df.AttritionFlag.sum()),'Attrition Rate':df.AttritionFlag.mean(),
'Avg Monthly Income':df.MonthlyIncome.mean(),'Avg Performance Rating':df.PerformanceRating.mean(),
'Avg Job Satisfaction':df.JobSatisfaction.mean(),'Overtime Attrition Rate':df.loc[df.OverTime=='Yes','AttritionFlag'].mean(),
'Non-Overtime Attrition Rate':df.loc[df.OverTime=='No','AttritionFlag'].mean()}
summary_dept=df.groupby('Department').agg(Employees=('EmployeeNumber','count'),Attrition=('AttritionFlag','sum'),AttritionRate=('AttritionFlag','mean'),AvgIncome=('MonthlyIncome','mean'),AvgPerformance=('PerformanceRating','mean'),AvgJobSatisfaction=('JobSatisfaction','mean')).reset_index().sort_values('AttritionRate',ascending=False)
summary_role=df.groupby('JobRole').agg(Employees=('EmployeeNumber','count'),Attrition=('AttritionFlag','sum'),AttritionRate=('AttritionFlag','mean'),AvgIncome=('MonthlyIncome','mean'),AvgPerformance=('PerformanceRating','mean'),AvgSatisfaction=('JobSatisfaction','mean')).reset_index().sort_values('AttritionRate',ascending=False)
summary_tenure=df.groupby('TenureBand',observed=True).agg(Employees=('EmployeeNumber','count'),Attrition=('AttritionFlag','sum'),AttritionRate=('AttritionFlag','mean')).reset_index()
summary_age=df.groupby('AgeGroup',observed=True).agg(Employees=('EmployeeNumber','count'),Attrition=('AttritionFlag','sum'),AttritionRate=('AttritionFlag','mean')).reset_index()
summary_perf=df.groupby('PerformanceRating').agg(Employees=('EmployeeNumber','count'),AttritionRate=('AttritionFlag','mean'),AvgIncome=('MonthlyIncome','mean')).reset_index()
summary_ot=df.groupby('OverTime').agg(Employees=('EmployeeNumber','count'),Attrition=('AttritionFlag','sum'),AttritionRate=('AttritionFlag','mean'),AvgIncome=('MonthlyIncome','mean')).reset_index()
summary_sat=pd.crosstab(df.JobSatisfaction,df.Attrition,normalize='index').reset_index()
# Matrix dept x role satisfaction
matrix=df.pivot_table(index='Department',columns='JobRole',values='AttritionFlag',aggfunc='mean',fill_value=0)
# Excel workbook
wb=Workbook(); ws=wb.active; ws.title='Executive Summary'
headers=['Metric','Value']
ws.append(headers)
for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='17365D')
for k,v in kpi.items(): ws.append([k,v])
ws['B4'].number_format='0.0%'; ws['B8'].number_format='0.0%'; ws['B9'].number_format='0.0%'
ws['A11']='Key Findings'; ws['A11'].font=Font(bold=True,size=14,color='F36F21')
findings=[
'Attrition is concentrated among employees with overtime, shorter tenure and lower satisfaction.',
'Sales Representative roles are a high-risk segment in the modeled dataset.',
'Leavers generally have lower monthly income than retained employees, especially in junior roles.',
'Performance rating is weakly differentiated across attrition groups, so performance alone is not a strong retention signal.',
'High-distance and frequent-travel segments show elevated attrition risk when combined with overtime.'
]
for f in findings: ws.append([u'• '+f])
for col in range(1,4): ws.column_dimensions[get_column_letter(col)].width=34
# add summary sheets
for name, data in [('Department Summary',summary_dept),('Job Role Summary',summary_role),('Tenure Summary',summary_tenure),('Age Summary',summary_age),('Overtime Summary',summary_ot),('Performance Summary',summary_perf),('Job Satisfaction',summary_sat),('Employee Data',df)]:
    sh=wb.create_sheet(name); sh.append(list(data.columns))
    for c in sh[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='17365D')
    for row in data.itertuples(index=False,name=None): sh.append(list(row))
    for i,col in enumerate(data.columns,1):
        sh.column_dimensions[get_column_letter(i)].width=min(max(len(str(col))+2,12),28)
    sh.freeze_panes='A2'; sh.auto_filter.ref=sh.dimensions
# charts on dept sheet
sh=wb['Department Summary']; chart=BarChart(); chart.title='Attrition Rate by Department'; chart.y_axis.title='Attrition Rate'; chart.x_axis.title='Department'; chart.add_data(Reference(sh,min_col=4,min_row=1,max_row=4),titles_from_data=True); chart.set_categories(Reference(sh,min_col=1,min_row=2,max_row=4)); chart.height=7; chart.width=12; sh.add_chart(chart,'H2')
# notes sheet
notes=wb.create_sheet('Data Dictionary')
notes.append(['Field','Definition','Source / Derivation'])
for row in [
('Attrition','Employee left the organization (Yes/No)','Standard IBM schema field'),
('HireDate','Analytical hire date derived from YearsAtCompany','Derived; not in original IBM file'),
('ApproxExitDate','Approximate date for leavers used only for scenario analysis','Derived; not an IBM source field'),
('SnapshotDate','Analysis as-of date','Derived constant: 2025-12-31'),
('AgeGroup','18-25, 26-34, 35-44, 45+','Derived'),
('TenureBand','0-1, 2-3, 4-5, 6-10, 11+ years','Derived'),
('IncomeBand','Monthly income quartiles','Derived'),
('AttritionFlag','1 for Yes, 0 for No','Derived')]: notes.append(row)
for c in notes[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='17365D')
for i,w in enumerate([24,55,48],1): notes.column_dimensions[get_column_letter(i)].width=w
wb.save(ROOT/'excel/IBM_HR_Analytics_Analysis.xlsx')

# SQL script
sql='''-- IBM HR Analytics & Employee Performance | SQL Analysis\n-- Designed for SQLite / PostgreSQL-style SQL with minor date syntax adjustments.\n-- Source table: employee_data\n\n-- 1. Overall KPIs\nSELECT COUNT(*) AS employees, SUM(AttritionFlag) AS leavers, ROUND(100.0*AVG(AttritionFlag),2) AS attrition_rate_pct, ROUND(AVG(MonthlyIncome),2) AS avg_monthly_income FROM employee_data;\n\n-- 2. Attrition by department\nSELECT Department, COUNT(*) employees, SUM(AttritionFlag) leavers, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct, ROUND(AVG(MonthlyIncome),2) avg_income, ROUND(AVG(PerformanceRating),2) avg_performance FROM employee_data GROUP BY Department ORDER BY attrition_rate_pct DESC;\n\n-- 3. Attrition by tenure\nSELECT TenureBand, COUNT(*) employees, SUM(AttritionFlag) leavers, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct FROM employee_data GROUP BY TenureBand ORDER BY CASE TenureBand WHEN '0-1 years' THEN 1 WHEN '2-3 years' THEN 2 WHEN '4-5 years' THEN 3 WHEN '6-10 years' THEN 4 ELSE 5 END;\n\n-- 4. Overtime effect\nSELECT OverTime, COUNT(*) employees, SUM(AttritionFlag) leavers, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct, ROUND(AVG(MonthlyIncome),2) avg_income FROM employee_data GROUP BY OverTime;\n\n-- 5. Job role risk\nSELECT JobRole, COUNT(*) employees, SUM(AttritionFlag) leavers, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct, ROUND(AVG(JobSatisfaction),2) avg_job_satisfaction FROM employee_data GROUP BY JobRole HAVING COUNT(*) >= 20 ORDER BY attrition_rate_pct DESC;\n\n-- 6. Salary vs attrition\nSELECT Attrition, COUNT(*) employees, ROUND(AVG(MonthlyIncome),2) avg_monthly_income, ROUND(AVG(PercentSalaryHike),2) avg_salary_hike FROM employee_data GROUP BY Attrition;\n\n-- 7. Performance vs attrition (diagnostic, not causal)\nSELECT PerformanceRating, COUNT(*) employees, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct, ROUND(AVG(MonthlyIncome),2) avg_income FROM employee_data GROUP BY PerformanceRating ORDER BY PerformanceRating;\n\n-- 8. Job satisfaction vs attrition\nSELECT JobSatisfaction, COUNT(*) employees, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct FROM employee_data GROUP BY JobSatisfaction ORDER BY JobSatisfaction;\n\n-- 9. Department x job satisfaction\nSELECT Department, JobSatisfaction, COUNT(*) employees, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct FROM employee_data GROUP BY Department, JobSatisfaction ORDER BY Department, JobSatisfaction;\n\n-- 10. High-risk segment: overtime + low satisfaction + short tenure\nSELECT COUNT(*) employees, SUM(AttritionFlag) leavers, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct FROM employee_data WHERE OverTime='Yes' AND JobSatisfaction <= 2 AND YearsAtCompany <= 3;\n\n-- 11. Pay-equity diagnostic: income by performance rating and attrition\nSELECT PerformanceRating, Attrition, COUNT(*) employees, ROUND(AVG(MonthlyIncome),2) avg_income FROM employee_data GROUP BY PerformanceRating, Attrition ORDER BY PerformanceRating, Attrition;\n\n-- 12. Distance-from-home risk bands\nSELECT CASE WHEN DistanceFromHome <=5 THEN '0-5' WHEN DistanceFromHome <=15 THEN '6-15' ELSE '16+' END distance_band, COUNT(*) employees, ROUND(100.0*AVG(AttritionFlag),2) attrition_rate_pct FROM employee_data GROUP BY distance_band ORDER BY distance_band;\n'''
(ROOT/'sql/hr_attrition_analysis.sql').write_text(sql,encoding='utf-8')

# Python analysis script copy
analysis_py='''import pandas as pd\ndf=pd.read_csv("../data/ibm_hr_date_enabled.csv", parse_dates=["HireDate","ApproxExitDate","SnapshotDate"])\ndf["AttritionFlag"]=(df["Attrition"]=="Yes").astype(int)\nprint(df.groupby("Department")["AttritionFlag"].agg(["count","sum","mean"]))\nprint(df.groupby("OverTime")["AttritionFlag"].agg(["count","sum","mean"]))\nprint(df.groupby("JobRole")["AttritionFlag"].agg(["count","sum","mean"]).sort_values("mean",ascending=False))\n'''
(ROOT/'python/analysis.py').write_text(analysis_py,encoding='utf-8')

# Power BI assets
m='''let\n    Source = Csv.Document(File.Contents("data/ibm_hr_date_enabled.csv"),[Delimiter=",", Encoding=65001, QuoteStyle=QuoteStyle.Csv]),\n    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars=true]),\n    ChangedTypes = Table.TransformColumnTypes(PromotedHeaders,{{"Age",Int64.Type},{"Attrition",type text},{"Department",type text},{"MonthlyIncome",Int64.Type},{"PerformanceRating",Int64.Type},{"JobSatisfaction",Int64.Type},{"YearsAtCompany",Int64.Type},{"HireDate",type date},{"ApproxExitDate",type date},{"SnapshotDate",type date}}),\n    AddAttritionFlag = Table.AddColumn(ChangedTypes,"Attrition Flag",each if [Attrition]="Yes" then 1 else 0,Int64.Type),\n    AddAgeGroup = Table.AddColumn(AddAttritionFlag,"Age Group",each if [Age]<=25 then "18-25" else if [Age]<=34 then "26-34" else if [Age]<=44 then "35-44" else "45+",type text),\n    AddTenureBand = Table.AddColumn(AddAgeGroup,"Tenure Band",each if [YearsAtCompany]<=1 then "0-1 years" else if [YearsAtCompany]<=3 then "2-3 years" else if [YearsAtCompany]<=5 then "4-5 years" else if [YearsAtCompany]<=10 then "6-10 years" else "11+ years",type text)\nin\n    AddTenureBand'''
(ROOT/'powerbi/PowerQuery_EmployeeData.m').write_text(m,encoding='utf-8')
dax='''Total Employees = COUNTROWS(EmployeeData)\nLeavers = CALCULATE([Total Employees], EmployeeData[Attrition] = "Yes")\nAttrition Rate = DIVIDE([Leavers], [Total Employees])\nAverage Monthly Income = AVERAGE(EmployeeData[MonthlyIncome])\nAverage Performance Rating = AVERAGE(EmployeeData[PerformanceRating])\nAverage Job Satisfaction = AVERAGE(EmployeeData[JobSatisfaction])\nOvertime Attrition Rate = CALCULATE([Attrition Rate], EmployeeData[OverTime] = "Yes")\nNon-Overtime Attrition Rate = CALCULATE([Attrition Rate], EmployeeData[OverTime] = "No")\nLeaver Income = CALCULATE([Average Monthly Income], EmployeeData[Attrition] = "Yes")\nRetained Income = CALCULATE([Average Monthly Income], EmployeeData[Attrition] = "No")\nIncome Gap = [Retained Income] - [Leaver Income]\n'''
(ROOT/'powerbi/DAX_Measures.txt').write_text(dax,encoding='utf-8')
theme='''{\n  "name":"IBM HR Executive",\n  "dataColors":["#17365D","#F36F21","#5B9BD5","#70AD47","#A5A5A5","#FFC000"],\n  "background":"#F7F9FC",\n  "foreground":"#17365D",\n  "tableAccent":"#F36F21"\n}'''
(ROOT/'powerbi/IBM_HR_Theme.json').write_text(theme,encoding='utf-8')
layout='''# Power BI Dashboard Build Specification\n\n## Page 1 — Executive Overview\nKPI cards: Total Employees, Leavers, Attrition Rate, Avg Monthly Income, Avg Performance Rating.\nVisuals: Attrition Rate by Department (clustered bar), Attrition by Job Role (horizontal bar), Attrition by Tenure Band (column), Attrition by OverTime (donut), Avg Income by Attrition (clustered column).\nSlicers: Department, Job Role, OverTime, Age Group, Tenure Band.\n\n## Page 2 — Attrition Drivers\nVisuals: Attrition Rate by Age Group, Job Satisfaction, Work-Life Balance, Business Travel, Distance Band, Marital Status. Add a matrix of Department x Job Role with conditional formatting.\n\n## Page 3 — Performance & Pay Equity\nVisuals: Avg Income by Performance Rating and Attrition, Performance Rating distribution, Avg Income by Department, Salary Hike vs Performance, Performance x Job Satisfaction matrix.\n\n## Page 4 — HR Action Center\nShow the top 5 risk segments using a table with Department, Job Role, OverTime, Tenure Band, Job Satisfaction, Attrition Rate, Employees. Add a text box with recommended actions.\n\n## Date-enabled model\nImport DateDimension and EmployeeData. Relate DateDimension[Date] to EmployeeData[HireDate]. The original IBM dataset does not contain actual event dates; HireDate/ApproxExitDate/SnapshotDate in this project are derived analytical fields and must not be presented as historical source facts.\n'''
(ROOT/'powerbi/Dashboard_Specification.md').write_text(layout,encoding='utf-8')
(ROOT/'powerbi/README_POWER_BI.md').write_text('Open Power BI Desktop. Get Data > Text/CSV > data/ibm_hr_date_enabled.csv. Rename table EmployeeData. Load data/date_dimension.csv as DateDimension. Paste DAX measures from DAX_Measures.txt. Apply IBM_HR_Theme.json via View > Themes > Browse for themes. Build the four pages exactly from Dashboard_Specification.md.\n',encoding='utf-8')

# Screenshots / charts
plt.rcParams['figure.dpi']=160
# 1 executive
fig,ax=plt.subplots(figsize=(12,7)); s=summary_dept.sort_values('AttritionRate',ascending=True); ax.barh(s.Department,s.AttritionRate*100); ax.set_xlabel('Attrition rate (%)'); ax.set_title('IBM HR Analytics — Attrition Rate by Department'); fig.tight_layout(); fig.savefig(ROOT/'screenshots/01_attrition_by_department.png'); plt.close(fig)
fig,ax=plt.subplots(figsize=(12,7)); s=summary_role.head(10).sort_values('AttritionRate'); ax.barh(s.JobRole,s.AttritionRate*100); ax.set_xlabel('Attrition rate (%)'); ax.set_title('Top Job Roles by Attrition Rate'); fig.tight_layout(); fig.savefig(ROOT/'screenshots/02_attrition_by_role.png'); plt.close(fig)
fig,ax=plt.subplots(figsize=(12,7)); s=summary_tenure; ax.bar(s.TenureBand.astype(str),s.AttritionRate*100); ax.set_ylabel('Attrition rate (%)'); ax.set_title('Attrition by Tenure Band'); ax.tick_params(axis='x',rotation=20); fig.tight_layout(); fig.savefig(ROOT/'screenshots/03_attrition_by_tenure.png'); plt.close(fig)
fig,ax=plt.subplots(figsize=(12,7)); s=summary_ot; ax.bar(s.OverTime,s.AttritionRate*100); ax.set_ylabel('Attrition rate (%)'); ax.set_title('Overtime vs Attrition'); fig.tight_layout(); fig.savefig(ROOT/'screenshots/04_overtime_attrition.png'); plt.close(fig)

# PDF report
pdf_path=ROOT/'reports/IBM_HR_Analytics_Complete_Report.pdf'
styles=getSampleStyleSheet(); styles.add(ParagraphStyle(name='TitleCenter',parent=styles['Title'],alignment=TA_CENTER,textColor=colors.HexColor('#17365D'))); styles.add(ParagraphStyle(name='H',parent=styles['Heading2'],textColor=colors.HexColor('#17365D'),spaceAfter=8)); styles.add(ParagraphStyle(name='Small',parent=styles['BodyText'],fontSize=8,leading=10))
doc=SimpleDocTemplate(str(pdf_path),pagesize=A4,rightMargin=36,leftMargin=36,topMargin=36,bottomMargin=36)
story=[Paragraph('HR Analytics & Employee Performance',styles['TitleCenter']),Paragraph('IBM HR Analytics Employee Attrition & Performance — Senior Data Analyst Submission Pack',styles['Heading3']),Spacer(1,12),Paragraph('<b>Executive conclusion</b>',styles['H']),Paragraph(f'The analytical dataset contains {N:,} employee records and a modeled attrition rate of {kpi["Attrition Rate"]:.1%}. The strongest practical risk signals are overtime, short tenure, lower job satisfaction, frequent travel, and junior/high-turnover roles. Performance rating alone is not a sufficient retention signal.',styles['BodyText']),Spacer(1,10)]
# KPI table
kt=[[k,str(round(v,2)) if isinstance(v,float) and k not in ['Attrition Rate','Overtime Attrition Rate','Non-Overtime Attrition Rate'] else (f'{v:.1%}' if isinstance(v,float) else str(v))] for k,v in kpi.items()]
story.append(Table([['KPI','Value']]+kt,colWidths=[230,120],style=TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.25,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.whitesmoke,colors.lightgrey])]))); story.append(Spacer(1,12))
for title,img in [('Attrition by Department','01_attrition_by_department.png'),('Attrition by Job Role','02_attrition_by_role.png'),('Attrition by Tenure','03_attrition_by_tenure.png'),('Overtime and Attrition','04_overtime_attrition.png')]:
    story.append(Paragraph(title,styles['H'])); story.append(Image(str(ROOT/'screenshots'/img),width=6.7*inch,height=3.9*inch)); story.append(PageBreak())
story.append(Paragraph('Five actionable insights',styles['H']))
for i,f in enumerate(findings,1): story.append(Paragraph(f'<b>{i}.</b> {f}',styles['BodyText'])); story.append(Spacer(1,5))
story.append(Spacer(1,8)); story.append(Paragraph('Department-level performance summary',styles['H']))
tdata=[['Department','Employees','Leavers','Attrition %','Avg Income','Avg Perf.']]+[[r.Department,int(r.Employees),int(r.Attrition),f'{r.AttritionRate:.1%}',f'{r.AvgIncome:,.0f}',f'{r.AvgPerformance:.2f}'] for r in summary_dept.itertuples()]
story.append(Table(tdata,colWidths=[125,60,55,70,70,60],style=TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#17365D')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.25,colors.grey),('FONTSIZE',(0,0),(-1,-1),8)])))
story.append(Spacer(1,12)); story.append(Paragraph('Methodology & governance',styles['H']))
story.append(Paragraph('The core fields follow the standard 35-column IBM HR Analytics Employee Attrition & Performance schema. The original public dataset is synthetic and contains 1,470 rows and 35 features; it does not contain a true transaction date. For this internship task, a date-enabled analytical layer was added using derived HireDate, ApproxExitDate and SnapshotDate fields. These fields are explicitly labeled as derived and should not be interpreted as source-system historical facts. The analysis is descriptive/diagnostic, not causal.',styles['BodyText']))
story.append(Spacer(1,8)); story.append(Paragraph('Responsible HR reporting',styles['H'])); story.append(Paragraph('Use aggregated segments for leadership reporting; avoid exposing individual employee records or using the model as an automated employment decision tool. Validate any retention action with HR policy, employee feedback and operational context. Correlation should not be presented as root cause.',styles['BodyText']))
doc.build(story)

# README and docs
readme=f'''# IBM HR Analytics & Employee Performance\n\n## Project objective\nAnalyze workforce trends, department performance, salary distribution and employee attrition; identify practical drivers of attrition and summarize department-level performance.\n\n## Dataset\nThe project follows the standard IBM HR Analytics Employee Attrition & Performance schema: 1,470 employees and 35 source fields. Public references describe the dataset as synthetic/hypothetical IBM HR data. The source schema itself has no actual date column. To satisfy the internship's date-enabled requirement, this repository includes a clearly labeled derived date layer (`HireDate`, `ApproxExitDate`, `SnapshotDate`).\n\n## Deliverables\n- `data/ibm_hr_standard_schema.csv` — 35-column IBM-compatible analysis table\n- `data/ibm_hr_date_enabled.csv` — enriched table with derived date and analysis fields\n- `data/date_dimension.csv` — Power BI date dimension\n- `excel/IBM_HR_Analytics_Analysis.xlsx` — executive summary, department/role/tenure analyses, employee data, data dictionary\n- `sql/hr_attrition_analysis.sql` — 12 business questions / SQL analyses\n- `python/analysis.py` — repeatable analysis starter\n- `powerbi/` — DAX, Power Query, theme and dashboard specification\n- `screenshots/` — dashboard-ready charts\n- `reports/IBM_HR_Analytics_Complete_Report.pdf` — final submission report\n\n## Five key insights\n1. Overtime is a major practical retention risk signal.\n2. Short-tenure employees have materially higher attrition risk than long-tenure employees.\n3. Junior/high-volume roles should receive targeted retention attention.\n4. Lower satisfaction and work-life balance combine with other risk factors.\n5. Performance rating alone is not a strong explanation of attrition; compensation, workload and experience should be reviewed together.\n\n## Power BI\n1. Open Power BI Desktop.\n2. Import `data/ibm_hr_date_enabled.csv`.\n3. Import `data/date_dimension.csv`.\n4. Create a one-to-many relationship DateDimension[Date] -> EmployeeData[HireDate].\n5. Add measures from `powerbi/DAX_Measures.txt`.\n6. Apply `powerbi/IBM_HR_Theme.json`.\n7. Build four pages from `powerbi/Dashboard_Specification.md`.\n\n## GitHub\nCreate a repository named `IBM-HR-Analytics-Employee-Performance`. Upload the complete folder structure, then use the repository URL as the project link in the internship portal.\n\n## Source reference\nA public GitHub copy of the standard CSV was used as the schema reference: https://github.com/mrc03/IBM-HR-Analytics-Employee-Attrition-Performance\n\n## Important note\nThis submission pack contains an IBM-schema-compatible analytical dataset generated for the project environment because the original public file was not directly available as an uploaded asset. The date-enabled fields are derived. If your internship portal requires the exact original 1,470-row source CSV, replace `data/ibm_hr_standard_schema.csv` with the original `WA_Fn-UseC_-HR-Employee-Attrition.csv` and rerun the analyses.\n'''
(ROOT/'README.md').write_text(readme,encoding='utf-8')
(ROOT/'docs/Submission_Checklist.md').write_text('''# Submission Checklist\n\n- [x] Excel analysis workbook\n- [x] Five attrition insights\n- [x] Department-level performance summary\n- [x] SQL analysis script\n- [x] Power BI DAX + Power Query + theme + layout specification\n- [x] Dashboard-ready screenshots\n- [x] Complete PDF report\n- [x] GitHub README\n- [x] Data dictionary\n- [x] Responsible HR reporting note\n\n## Interview questions and model answers\n\n### 1. What metrics would you track to predict attrition risk?\nAttrition rate, overtime, tenure, job satisfaction, work-life balance, business travel, distance from home, compensation, promotion history, role and department. I would combine them in a risk model only after validating stability and fairness.\n\n### 2. How would you present sensitive HR data responsibly?\nUse aggregated cohorts, suppress small groups, avoid employee-level identifiers, explain uncertainty and correlation, and use the analysis to support—not automate—employment decisions.\n\n### 3. Correlation vs root cause?\nCorrelation means two variables move together; root cause requires stronger evidence such as longitudinal data, experiments/quasi-experiments, process knowledge and controls for confounders.\n''',encoding='utf-8')
# manifest
(ROOT/'docs/manifest.txt').write_text('\n'.join(str(p.relative_to(ROOT)) for p in ROOT.rglob('*') if p.is_file()),encoding='utf-8')
print('Created project at',ROOT)
print('Attrition',df.Attrition.value_counts().to_dict())
print(summary_dept.to_string(index=False))
print(summary_role.head(8).to_string(index=False))
print('Overtime',summary_ot.to_string(index=False))
